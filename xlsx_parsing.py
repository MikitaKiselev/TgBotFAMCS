from openpyxl import load_workbook

book = load_workbook(filename="D:/Downloaded_from_Telegram/1_course.xlsx")

sheet = book["Лист1"]

schedule = ""

letters = [["C", "D"], ["E", "F"], ["G", "H"], ["I", "J"], ["L", "M"], ["N", "O"], ["P", "Q"], ["S", "T"], ["V", "W"],
           ["X", "Y"], ["Z", "AA"], ["AB", "AC"]]

lecture1 = [13, 17, 34, 50, 55, 59, 67, 76, 88, 92, 97, 130]
lecture2 = [21, 25, 34, 42, 50, 55, 76, 88, 92, 105, 109, 126]
#lecture3 = [20, 16, 33, 54, 79, 83, 104, 108, 121]
lecture4 = [13, 25, 34, 46, 50, 63, 80, 92, 101]

xlsx_str = [12, 33, 54, 75, 96, 117, 137]

for groups in range(12):

    for i in range(7, 134):
        value_A = sheet["A" + str(i)].value
        value_B = sheet["B" + str(i)].value
        value_C = sheet[letters[groups][0] + str(i)].value
        value_D = sheet[letters[groups][1] + str(i)].value

        if groups >= 1 and groups <= 3:
            for val in lecture1:
                if i == val:
                    value_C = sheet["C" + str(i)].value

        elif groups >= 4 and groups <= 6:
            for val in lecture2:
                if i == val:
                    value_C = sheet["L" + str(i)].value

        elif groups >= 8 and groups <= 11:
            for val in lecture4:
                if i == val:
                    value_C = sheet["V" + str(i)].value

        for k in range(6):
            if i == xlsx_str[k]:
                schedule += "\n\n"

        if value_A is not None:
            schedule += str(value_A)
            schedule += "\n"

        if value_B is not None:
            schedule += str(value_B)
            schedule += " "

        if value_C is not None:
            if not sheet[letters[groups][0] + str(i)].font.italic:
                schedule += str(value_C)
                schedule += "\n"

        if value_D is not None:
            if not sheet[letters[groups][1] + str(i)].font.italic:
                schedule += str(value_D)
                schedule += "\n"



# print(schedule)

# start_cell = None
# for col in sheet.iter_cols():
#     for cell in col:
#         if cell.value == "2 группа":
#             start_cell = cell
#             break
#     if start_cell:
#         break
#
# # Переменная для хранения данных
# data = {}
# col_data = ""
# # Проходимся по столбцам с данными ниже "1 группа" и собираем данные
# for col in sheet.iter_cols(min_col=start_cell.column, min_row=start_cell.row + 1):
#
#     for cell in col:
#         if cell.value is not None:
#             col_data += str(cell.value) + "\n"
#     #data[col[0].column] = col_data
#
# # В переменной data теперь содержатся данные из всех столбцов ниже "1 группа"
# print(col_data)
file_path = "shedule1.txt"

# --------ЗАПИСЬ---------
# Открываем файл в режиме записи (w - write)
with open(file_path, 'w', encoding='utf-8') as file:
    file.write(schedule)



